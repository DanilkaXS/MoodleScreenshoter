from pprint import pprint

from openpyxl.reader.excel import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By

import gspread
from google.oauth2 import service_account
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from gspread.utils import ExportFormat
import openpyxl

import time

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service_account.json'
credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)
service = build('drive', 'v3', credentials=credentials, static_discovery=False)
sa = gspread.service_account('service_account.json')
# results = service.files().list(pageSize=10,
#                                fields="nextPageToken, files(id,name,description)").execute()


def get_data_from_gs():
    try:
        gsheet = "Довідка ІДГУ 2022-2023 (Відповіді)"
        sh = sa.open(f"{gsheet}")
        export_file = sh.export(format=ExportFormat.EXCEL)
        f = open('data/data.xlsx', 'wb')
        f.write(export_file)
        f.close()
        wb = load_workbook(filename="data/data.xlsx")
        raw_data_list = wb['Відповіді форми (1)']
        max_rows = wb['Відповіді форми (1)'].max_row

        def iter_rows(ws):
            for row in ws.iter_rows():
                yield [cell.value for cell in row]

        raw_data_list = list(iter_rows(raw_data_list))
        data = []

        for raw_data in raw_data_list:
            if raw_data[1] != None and raw_data[2] != 'Назва курсу ':
                data.append({"name": f"{raw_data[1]}",
                             "course": f"{raw_data[2]}",
                             "url": f"{raw_data[3]}",
                             "lecture": f"{raw_data[5]}",
                             "seminar": f"{raw_data[6]}",
                             "pract": f"{raw_data[7]}",
                             "lab": f"{raw_data[8]}",
                             "control": f"{raw_data[9]}",
                             }, )
            else:
                continue
        return data

    except Exception as _ex:
        print(_ex)


def open_web_page(url, username, password, driver):
    title = 1
    try:
        driver.get(url)
        time.sleep(3)
        # title = driver.title.split(':')[1]
        title = 1
        driver.get_full_page_screenshot_as_file(filename=f"screens/{title}.png")

    except Exception as _ex:
        print(f"[ERROR] f{_ex}")

    return title


def create_table(data, username, password, driver):
    name = data["name"]
    title = data["course"]
    url = data["url"]
    document = f"{name.split()[0]}_{title.split()[0]}"
    wb = Workbook()
    ws = wb.active
    # ws = wb.create_sheet(title=document)
    print(f"[INFO] Creating document for {name}")

    # Header
    ws.merge_cells('A1:H1')
    ws['A1'] = "Довідка про наповнення курсу в системі Moodle ІДГУ"
    header = ws['A1']
    header.alignment = Alignment(horizontal="center", vertical="center")
    header.font = Font(name='Times New Roman', size=14, bold=True)
    print(f"[INFO] Editing document {document}")
    # Table with data
    ws.merge_cells('A3:B3')
    ws['A3'] = "Назва курсу"
    ws.merge_cells('C3:H3')
    ws['C3'] = f"{title}"

    ws.merge_cells('A4:B4')
    ws['A4'] = "Прізвище та ім’я автора курсу"
    ws.merge_cells('C4:H4')
    ws['C4'] = f"{name}"

    ws.merge_cells('A5:B5')
    ws['A5'] = "Адреса розміщення"
    ws.merge_cells('C5:H5')
    ws['C5'] = f"{url}"

    ws.merge_cells('A6:H6')

    ws.merge_cells('A7:B7')
    ws['A7'] = "Лекції(кількість)"
    ws['C7'] = f"{data['lecture']}"
    ws.merge_cells('D7:H7')
    ws['D7'] = "год."

    ws.merge_cells('A8:B8')
    ws['A8'] = "Плани семінарських (кількість)"
    ws['C8'] = f"{data['seminar']}"
    ws.merge_cells('D8:H8')
    ws['D8'] = "год."

    ws.merge_cells('A9:B9')
    ws['A9'] = "Практичних"
    ws['C9'] = f"{data['pract']}"
    ws.merge_cells('D9:H9')
    ws['D9'] = "год."

    ws.merge_cells('A10:B10')
    ws['A10'] = "Лабораторних"
    ws['C10'] = f"{data['lab']}"
    ws.merge_cells('D10:H10')
    ws['D10'] = "год."

    ws.merge_cells('A11:B11')
    ws['A11'] = "Форма контролю"
    ws['C11'] = f"{data['control']}"
    ws.merge_cells('D11:H11')
    ws['D11'] = "год."

    def __format_ws__(ws, cell_range, size_text):
        # applying border and alignment
        font = Font(name='Times New Roman', size=size_text)
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        rows = [rows for rows in ws[cell_range]]
        flattened = [item for sublist in rows for item in sublist]
        [(setattr(cell, 'border', border), setattr(cell, 'font', font),) for cell in flattened]

    __format_ws__(ws=ws, cell_range='A3:H11',
                  size_text=12)

    edit = ws["C3"]
    edit.alignment = Alignment(horizontal="center", vertical="center")
    edit.font = Font(name='Times New Roman', size=11)
    edit = ws["C4"]
    edit.alignment = Alignment(horizontal="center", vertical="center")
    edit.font = Font(name='Times New Roman', size=12)

    # Header 2
    ws.merge_cells('A12:H12')
    ws['A12'] = "Резюме про наповнення курсу"
    header2 = ws['A12']
    header2.alignment = Alignment(horizontal="center", vertical="center")
    header2.font = Font(name='Times New Roman', size=14, bold=True)

    # Field 1

    ws.merge_cells('A13:H13')
    ws["A13"] = "Курс наповнено різноманітними матеріалами, додатковими ресурсами, засобами комунікації зі " \
                "студентами, засобами встановлення рівня якості з окремих тем і з курсу взагалі"

    ws.merge_cells('A15:H15')
    ws["A13"].font = Font(name='Times New Roman', size=12)
    ws['A13'].alignment = Alignment(wrap_text=True)

    ws[
        "A15"] = "Містить лише елементи зазначені у навантажені ( лекція, практичні/лабораторні/семінар/, МКР, " \
                 "екзамен) без використання можливостей системи"
    ws["A15"].font = Font(name='Times New Roman', size=12, bold=True, underline="single")
    ws['A15'].alignment = Alignment(wrap_text=True)

    ws.merge_cells('A17:H17')
    ws["A17"] = "Курс наповнено хаотично, обсяг матеріалів і завдань мінімальний. Рекомендується суттєве доопрацювання"
    ws["A17"].font = Font(name='Times New Roman', size=12)
    ws['A17'].alignment = Alignment(wrap_text=True)

    # Admin sign
    ws.merge_cells('A21:C21')
    ws['A21'] = "Підпис адміністратора системи"
    text = ws['A21']
    text.alignment = Alignment(horizontal="center", vertical="center")
    text.font = Font(name='Times New Roman', size=14)

    ws.merge_cells('D21:H21')
    ws['D21'] = "__________________________________________________"
    field = ws['D21']
    field.alignment = Alignment(horizontal="center", vertical="center")

    # Text img
    ws['E23'] = "Додаток: скрин зазначеного курсу"
    edit = ws["E23"]
    edit.font = Font(name='Times New Roman', size=12)

    # Screenshot
    print(f"[INFO] Getting screenshot for {title}")
    title = open_web_page(url, username, password, driver)

    img = openpyxl.drawing.image.Image(f"screens/{title}.png")
    # img.height = 1511.8112140865
    # img.width = 566.92920528242
    img.anchor = 'A51'
    ws.add_image(img)

    ws.column_dimensions['B'].width = 23
    ws.column_dimensions['F'].width = 14

    ws.row_dimensions[13].height = 32
    ws.row_dimensions[15].height = 32
    ws.row_dimensions[17].height = 32

    # Save doc
    print(f"[INFO] Saving {document}.xlsx ...")
    wb.save(filename=f"tables/{document}.xlsx")


def main():
    username = input("[INPUT] Login:")
    password = input("[INPUT] Password:")
    try:
        service = Service(executable_path="firefoxdriver/geckodriver.exe")
        options = webdriver.FirefoxOptions()
        options.add_argument('-headless')
        options.binary_location = r'C:\Program Files\Mozilla Firefox\firefox.exe'
        driver = webdriver.Firefox(service=service, options=options)
        driver.get("http://moodle.idgu.edu.ua/moodle/login/index.php")
        time.sleep(2)
        driver.find_element(By.ID, "username").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.ID, "loginbtn").click()
        print(f"[INFO] Successfully connected {username}")

        data = get_data_from_gs()
        for i in data:
            create_table(data=i, username=username, password=password, driver=driver)

        driver.close()
        driver.quit()
        print("[END] Thanks for using >3\nProgram closing ...")
        exit()
    except Exception as _ex:
        print(f"[ERROR] {_ex}")


if __name__ == '__main__':
    main()
